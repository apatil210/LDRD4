from io import BytesIO
import re

import pandas as pd
import plotly.express as px
import plotly.io as pio
import requests
import streamlit as st
from openpyxl import load_workbook

# ----------------------------
# App configuration
# ----------------------------
st.set_page_config(
    page_title="US Manufacturing Energy Classification",
    layout="wide"
)

pio.templates.default = "plotly"

DATA_URL = "https://raw.githubusercontent.com/apatil210/LDRD3/main/DatasetJune25.xlsx"
SHEET_NAME = "Process-level data"

TEXT_COLOR = "#14212B"
PAPER_BG = "rgba(0,0,0,0)"
PLOT_BG = "rgba(0,0,0,0)"
BAR_COLOR = "#0B6E74"

ENERGY_COLOR_MAP = {
    "Annual Electricity": "#54A24B",
    "Annual Fuels": "#F58518",
    "Annual Steam": "#4C78A8",
}

TEMP_COLOR_MAP = {
    "<100 °C": "#72B7B2",
    "100-200 °C": "#54A24B",
    "200-400 °C": "#ECA82C",
    ">400 °C": "#E45756",
}

# ----------------------------
# Exact workbook column names
# ----------------------------
COL_L2 = "Unit operation (Level 2 classification)"
COL_L3 = "Industrial process"
COL_PERCENT_ENERGY = "Percent Annual energy demand in 2022"

COL_ANNUAL_PRODUCTION = "Annual production in 2022\n(based on FU)"

COL_SEC_ELECTRICITY = "SEC electricity"
COL_SEC_FUELS = "SEC fuels"
COL_SEC_STEAM = "SEC fuels or electricity for steam or steam from CHP"

COL_EFFICIENCY = "Efficiency"
COL_PROCESS_TEMP = "Process temperature"
COL_PROCESS_TEMP_WEB = "Process Temperature for Webpage"
COL_INLET_TEMP = "Inlet temperature"
COL_OUTLET_TEMP = "Outlet temperature"
COL_PROCESS_PRESSURE = "Process pressure"
COL_INLET_PRESSURE = "Inlet pressure"
COL_OUTLET_PRESSURE = "Outlet pressure"
COL_RESIDENCE_TIME = "Residence time"
COL_NAICS = "NAICS Level 2 Code Number"

# ----------------------------
# Exact Excel columns (1-based)
# ----------------------------
EXCEL_COL_E = 5    # Description
EXCEL_COL_K = 11   # Temperature for webpage
EXCEL_COL_AU = 47  # Annual energy
EXCEL_COL_AW = 49  # Annual electricity
EXCEL_COL_AX = 50  # Annual fuels
EXCEL_COL_AY = 51  # Annual steam

# ----------------------------
# Header utilities
# ----------------------------
def normalize_header(value) -> str:
    text = str(value if value is not None else "")
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def canonical_header(value) -> str:
    text = normalize_header(value).lower()
    text = text.replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def find_matching_column(df: pd.DataFrame, target: str) -> str:
    target_canonical = canonical_header(target)

    for col in df.columns:
        if canonical_header(col) == target_canonical:
            return col

    available = "\n".join([f"- {repr(col)}" for col in df.columns])
    raise KeyError(
        f"Could not find required column: {target}\nAvailable columns:\n{available}"
    )

# ----------------------------
# Cleaning utilities
# ----------------------------
def clean_category(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    )


def clean_naics(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    )


def clean_text(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    )


def to_numeric_safe(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace("\xa0", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce")

# ----------------------------
# Data loading
# ----------------------------
@st.cache_data(show_spinner=False)
def load_excel_data(url: str) -> pd.DataFrame:
    response = requests.get(url, timeout=30)
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "")
    if "text/html" in content_type.lower():
        raise ValueError("The URL returned HTML instead of an Excel file.")

    wb = load_workbook(filename=BytesIO(response.content), data_only=True)
    ws = wb[SHEET_NAME]

    header_row_excel = 2
    data_start_excel = 4

    max_col = ws.max_column
    max_row = ws.max_row

    headers = [
        ws.cell(row=header_row_excel, column=c).value
        for c in range(1, max_col + 1)
    ]

    rows = []
    for r in range(data_start_excel, max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in row_values):
            continue
        rows.append(row_values)

    df = pd.DataFrame(
        rows,
        columns=[str(h) if h is not None else f"Unnamed_{i+1}" for i, h in enumerate(headers)]
    )

    # Exact Excel-sourced columns
    df[COL_PROCESS_TEMP_WEB] = pd.Series([row[EXCEL_COL_K - 1] for row in rows])
    df["_Description_Excel_E"] = pd.Series([row[EXCEL_COL_E - 1] for row in rows])
    df["_Annual_Energy_AU"] = pd.Series([row[EXCEL_COL_AU - 1] for row in rows])
    df["_Annual_Electricity_AW"] = pd.Series([row[EXCEL_COL_AW - 1] for row in rows])
    df["_Annual_Fuels_AX"] = pd.Series([row[EXCEL_COL_AX - 1] for row in rows])
    df["_Annual_Steam_AY"] = pd.Series([row[EXCEL_COL_AY - 1] for row in rows])

    return df

# ----------------------------
# Data preparation
# ----------------------------
def prepare_bar_data(df: pd.DataFrame) -> pd.DataFrame:
    l2_col = find_matching_column(df, COL_L2)
    pct_col = find_matching_column(df, COL_PERCENT_ENERGY)

    working_df = df[[l2_col, pct_col]].copy()
    working_df[l2_col] = clean_category(working_df[l2_col])
    working_df[pct_col] = to_numeric_safe(working_df[pct_col]).fillna(0)

    grouped_df = (
        working_df.groupby(l2_col, as_index=False)[pct_col]
        .sum()
        .sort_values(pct_col, ascending=False)
        .reset_index(drop=True)
    )

    grouped_df = grouped_df[grouped_df[pct_col] != 0].copy()

    max_abs_val = grouped_df[pct_col].abs().max()
    if pd.notna(max_abs_val) and max_abs_val <= 1:
        grouped_df["Display Percent"] = grouped_df[pct_col] * 100
    else:
        grouped_df["Display Percent"] = grouped_df[pct_col]

    grouped_df = grouped_df.rename(columns={
        l2_col: "Unit operation (Level 2 classification)",
        pct_col: "Raw Percent Annual energy demand in 2022"
    })

    grouped_df["Rank"] = range(1, len(grouped_df) + 1)

    return grouped_df


def build_fact_sheet(df: pd.DataFrame, selected_l2: str):
    l2_col = find_matching_column(df, COL_L2)
    l3_col = find_matching_column(df, COL_L3)
    annual_prod_col = find_matching_column(df, COL_ANNUAL_PRODUCTION)

    sec_elec_col = find_matching_column(df, COL_SEC_ELECTRICITY)
    sec_fuels_col = find_matching_column(df, COL_SEC_FUELS)
    sec_steam_col = find_matching_column(df, COL_SEC_STEAM)

    efficiency_col = find_matching_column(df, COL_EFFICIENCY)
    process_temp_col = find_matching_column(df, COL_PROCESS_TEMP)
    inlet_temp_col = find_matching_column(df, COL_INLET_TEMP)
    outlet_temp_col = find_matching_column(df, COL_OUTLET_TEMP)
    process_pressure_col = find_matching_column(df, COL_PROCESS_PRESSURE)
    inlet_pressure_col = find_matching_column(df, COL_INLET_PRESSURE)
    outlet_pressure_col = find_matching_column(df, COL_OUTLET_PRESSURE)
    residence_time_col = find_matching_column(df, COL_RESIDENCE_TIME)
    naics_col = find_matching_column(df, COL_NAICS)

    annual_energy_col = "_Annual_Energy_AU"
    annual_elec_col = "_Annual_Electricity_AW"
    annual_fuels_col = "_Annual_Fuels_AX"
    annual_steam_col = "_Annual_Steam_AY"
    description_col = "_Description_Excel_E"
    process_temp_web_col = COL_PROCESS_TEMP_WEB

    fact_df = df.copy()
    fact_df[l2_col] = clean_category(fact_df[l2_col])
    fact_df[l3_col] = clean_category(fact_df[l3_col])
    fact_df[naics_col] = clean_naics(fact_df[naics_col])

    selected_l2_clean = str(selected_l2).replace("\xa0", " ").strip()
    selected_df = fact_df[fact_df[l2_col] == selected_l2_clean].copy()

    if selected_df.empty:
        return None

    numeric_cols = [
        annual_prod_col,
        annual_energy_col,
        annual_elec_col,
        annual_fuels_col,
        annual_steam_col,
        sec_elec_col,
        sec_fuels_col,
        sec_steam_col,
        process_temp_col,
        process_temp_web_col,
    ]

    for col in numeric_cols:
        selected_df[col] = to_numeric_safe(selected_df[col])

    production_values = (
        selected_df[annual_prod_col]
        .dropna()
        .loc[lambda s: s != 0]
        .unique()
    )
    annual_production = production_values[0] if len(production_values) > 0 else 0

    annual_energy = selected_df[annual_energy_col].fillna(0).sum()
    annual_electricity = selected_df[annual_elec_col].fillna(0).sum()
    annual_fuels = selected_df[annual_fuels_col].fillna(0).sum()
    annual_steam = selected_df[annual_steam_col].fillna(0).sum()

    selected_df[description_col] = clean_text(selected_df[description_col])

    temp_energy_df = pd.DataFrame({
        "Temperature Raw": selected_df[process_temp_web_col],
        "Annual Energy AU": selected_df[annual_energy_col],
    }).dropna(subset=["Temperature Raw", "Annual Energy AU"])

    temp_energy_df["Annual Energy Magnitude"] = temp_energy_df["Annual Energy AU"].abs()
    temp_energy_df = temp_energy_df[temp_energy_df["Annual Energy Magnitude"] > 0].copy()

    if not temp_energy_df.empty:
        temp_energy_df["Temperature Level"] = pd.cut(
            temp_energy_df["Temperature Raw"],
            bins=[-float("inf"), 100, 200, 400, float("inf")],
            labels=["<100 °C", "100-200 °C", "200-400 °C", ">400 °C"],
            right=False
        )

        temp_breakdown_df = (
            temp_energy_df.groupby("Temperature Level", observed=False)["Annual Energy Magnitude"]
            .sum()
            .reset_index()
            .rename(columns={"Annual Energy Magnitude": "Value"})
        )

        temp_breakdown_df = temp_breakdown_df[temp_breakdown_df["Value"] > 0].copy()
    else:
        temp_breakdown_df = pd.DataFrame(columns=["Temperature Level", "Value"])

    detail_df = selected_df[
        [
            l3_col,
            description_col,
            naics_col,
            sec_elec_col,
            sec_fuels_col,
            sec_steam_col,
            efficiency_col,
            process_temp_col,
            process_temp_web_col,
            annual_energy_col,
            inlet_temp_col,
            outlet_temp_col,
            process_pressure_col,
            inlet_pressure_col,
            outlet_pressure_col,
            residence_time_col,
        ]
    ].rename(columns={
        l3_col: "List of Industry Application",
        description_col: "Description",
        naics_col: "NAICS Code",
        sec_elec_col: "SEC Electricity (GJ/t)",
        sec_fuels_col: "SEC Fuels (GJ/t)",
        sec_steam_col: "SEC Steam (GJ/t)",
        efficiency_col: "Efficiency (%)",
        process_temp_col: "Process temperature (°C)",
        process_temp_web_col: "Process Temperature for Webpage (°C)",
        annual_energy_col: "Annual Energy from AU",
        inlet_temp_col: "Inlet temperature (°C)",
        outlet_temp_col: "Outlet temperature (°C)",
        process_pressure_col: "Process pressure (bar)",
        inlet_pressure_col: "Inlet pressure (bar)",
        outlet_pressure_col: "Outlet pressure (bar)",
        residence_time_col: "Residence time (sec)",
    })

    return {
        "Annual Production": annual_production,
        "Annual Energy": annual_energy,
        "Annual Electricity": annual_electricity,
        "Annual Fuels": annual_fuels,
        "Annual Steam": annual_steam,
        "Temperature Breakdown": temp_breakdown_df,
        "Rows": len(selected_df),
        "Details": detail_df,
        "Temperature Source Column": process_temp_web_col,
        "Annual Energy Source Column": "AU",
    }

# ----------------------------
# Chart builders
# ----------------------------
def build_bar_chart(df: pd.DataFrame):
    fig = px.bar(
        df,
        x="Display Percent",
        y="Unit operation (Level 2 classification)",
        orientation="h",
        text="Display Percent",
        color_discrete_sequence=[BAR_COLOR]
    )

    fig.update_traces(
        texttemplate="%{text:.1f}%",
        textposition="outside",
        cliponaxis=False,
        hovertemplate="<b>%{y}</b><br>Summed percent annual energy: %{x:.4f}%<extra></extra>",
        marker=dict(line=dict(color="#FCFCFA", width=1.2))
    )

    fig.update_layout(
        width=1500,
        height=max(700, 34 * len(df)),
        paper_bgcolor=PAPER_BG,
        plot_bgcolor=PLOT_BG,
        margin=dict(t=60, l=340, r=80, b=30),
        xaxis_title="Percent Annual Energy Demand in 2022 (%)",
        yaxis_title="Unit Operation Classification",
        font=dict(family="Arial, sans-serif", color=TEXT_COLOR, size=14)
    )

    fig.update_xaxes(showgrid=True, automargin=True)
    fig.update_yaxes(categoryorder="total ascending", automargin=True, ticklabelstandoff=40)

    return fig


def build_annual_energy_donut(fact_sheet: dict):
    donut_df = pd.DataFrame({
        "Energy Type": ["Annual Electricity", "Annual Fuels", "Annual Steam"],
        "Value": [
            fact_sheet["Annual Electricity"],
            fact_sheet["Annual Fuels"],
            fact_sheet["Annual Steam"]
        ]
    })

    donut_df = donut_df[donut_df["Value"] > 0].copy()
    if donut_df.empty:
        return None

    fig = px.pie(
        donut_df,
        names="Energy Type",
        values="Value",
        hole=0.62,
        color="Energy Type",
        color_discrete_map=ENERGY_COLOR_MAP
    )

    total_energy = donut_df["Value"].sum()

    fig.update_traces(
        textposition="outside",
        texttemplate="%{label}<br>%{percent}",
        hovertemplate="<b>%{label}</b><br>Value: %{value:.3f} PJ/yr<br>Share: %{percent}<extra></extra>",
        marker=dict(line=dict(color="#FFFFFF", width=2))
    )

    fig.update_layout(
        height=360,
        margin=dict(t=20, l=20, r=20, b=20),
        paper_bgcolor=PAPER_BG,
        plot_bgcolor=PLOT_BG,
        showlegend=False,
        font=dict(family="Arial, sans-serif", color=TEXT_COLOR, size=13),
        annotations=[
            dict(
                text=f"<b>Total (PJ/yr)</b><br>{total_energy:.2f}",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=16, color=TEXT_COLOR)
            )
        ]
    )

    return fig


def build_temperature_donut(fact_sheet: dict):
    donut_df = fact_sheet["Temperature Breakdown"].copy()
    if donut_df.empty:
        return None

    fig = px.pie(
        donut_df,
        names="Temperature Level",
        values="Value",
        hole=0.62,
        color="Temperature Level",
        color_discrete_map=TEMP_COLOR_MAP
    )

    total_energy = donut_df["Value"].sum()

    fig.update_traces(
        textposition="outside",
        texttemplate="%{label}<br>%{percent}",
        hovertemplate="<b>%{label}</b><br>Energy magnitude from AU: %{value:.3f}<br>Share: %{percent}<extra></extra>",
        marker=dict(line=dict(color="#FFFFFF", width=2))
    )

    fig.update_layout(
        height=360,
        margin=dict(t=20, l=20, r=20, b=20),
        paper_bgcolor=PAPER_BG,
        plot_bgcolor=PLOT_BG,
        showlegend=False,
        font=dict(family="Arial, sans-serif", color=TEXT_COLOR, size=13),
        annotations=[
            dict(
                text=f"<b>Total (PJ/yr)</b><br>{total_energy:.2f}",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=16, color=TEXT_COLOR)
            )
        ]
    )

    return fig

# ----------------------------
# App UI
# ----------------------------
st.title("US Manufacturing Energy 2022 Classification: Unit Operations")

try:
    df = load_excel_data(DATA_URL)
    bar_df = prepare_bar_data(df)

    left_col, right_col = st.columns([1.6, 1.1], gap="large")

    with left_col:
        st.subheader("Percent Annual Energy by Unit Operation Classification")
        st.plotly_chart(
            build_bar_chart(bar_df),
            use_container_width=False,
            theme=None,
            config={"displayModeBar": False, "scrollZoom": False}
        )

    with right_col:
        selected_l2 = st.selectbox(
            "Select a unit operation (Level 2 classification) to generate a fact sheet",
            bar_df["Unit operation (Level 2 classification)"].tolist()
        )

        fact_sheet = build_fact_sheet(df, selected_l2)

        if fact_sheet is not None:
            st.subheader("Total Annual Energy Breakdown")

            st.caption("Categorization by Energy Source")
            donut_fig = build_annual_energy_donut(fact_sheet)
            if donut_fig is not None:
                st.plotly_chart(
                    donut_fig,
                    use_container_width=True,
                    theme=None,
                    config={"displayModeBar": False}
                )
            else:
                st.info("No positive annual energy values available for the selected category.")

            st.caption("Categorization by Process Temperature")
            temp_donut_fig = build_temperature_donut(fact_sheet)
            if temp_donut_fig is not None:
                st.plotly_chart(
                    temp_donut_fig,
                    use_container_width=True,
                    theme=None,
                    config={"displayModeBar": False}
                )
            else:
                st.info(
                    f"No nonzero annual energy magnitudes from column "
                    f"{fact_sheet['Annual Energy Source Column']} with valid "
                    f"'{fact_sheet['Temperature Source Column']}' temperatures are available for the selected category."
                )

            st.dataframe(
                fact_sheet["Details"].drop(columns=["Process Temperature for Webpage (°C)", "Annual Energy from AU"]),
                use_container_width=True,
                hide_index=True
            )

except Exception as e:
    st.error(f"App error: {e}")
